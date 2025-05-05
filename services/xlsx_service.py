import os
import re
from typing import Dict, Any, List, Optional
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment

from services.base_document_service import BaseDocumentService
from models.models import (
    Group, Student, Teacher, Discipline, Grade,
    ScheduleItem, Classroom
)


class XlsxService(BaseDocumentService):
    """
    Сервис для обработки XLSX-документов
    """

    async def load_document(self, template_path: str) -> openpyxl.Workbook:
        """
        Загружает XLSX-документ из файла шаблона

        Args:
            template_path: Путь к файлу шаблона

        Returns:
            Объект книги Excel
        """
        return openpyxl.load_workbook(template_path)

    async def save_document(self, document: openpyxl.Workbook, output_path: str) -> None:
        """
        Сохраняет XLSX-документ в файл

        Args:
            document: Объект книги Excel
            output_path: Путь для сохранения
        """
        document.save(output_path)

    async def fill_grades_journal(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, group_id: int,
                                  discipline_id: int, start_date: Optional[datetime] = None,
                                  period_weeks: int = 16) -> None:
        """
        Заполняет журнал оценок для группы по дисциплине

        Args:
            worksheet: Лист Excel
            group_id: ID группы
            discipline_id: ID дисциплины
            start_date: Начальная дата (по умолчанию текущая)
            period_weeks: Количество недель (по умолчанию 16)
        """

        group = await Group.get(id=group_id).prefetch_related('students')

        discipline = await Discipline.get(id=discipline_id).prefetch_related('control_works')

        if start_date is None:
            start_date = datetime.now()

        student_cells = []
        date_cells = []

        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value == 'N':
                    student_cells.append((row, col))
                elif cell.value == 'd':
                    date_cells.append((row, col))

        if not date_cells and not student_cells:

            header_row = next(
                (i for i in range(1, 11) if all(worksheet.cell(row=i, column=j).value for j in range(2, 6))), 1)

            student_col = 1
            for row in range(header_row + 1, min(header_row + 20, worksheet.max_row)):
                if worksheet.cell(row=row, column=student_col).value:
                    student_cells.append((row, student_col))

            for col in range(2, min(worksheet.max_column, 20)):
                date_cells.append((header_row, col))

        if not student_cells:

            for row in range(3, min(3 + len(group.students), 30)):
                student_cells.append((row, 1))

        if not date_cells:

            for col in range(2, min(2 + period_weeks, 20)):
                date_cells.append((1, col))

        if worksheet.cell(row=1, column=1).value:
            worksheet.cell(row=1, column=1).value = f"Журнал оценок - {discipline.name} - Группа {group.code}"

        for i, student in enumerate(group.students):
            if i < len(student_cells):
                row, col = student_cells[i]
                cell = worksheet.cell(row=row, column=col)
                cell.value = student.full_name

                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')

        current_date = start_date
        for i in range(min(period_weeks, len(date_cells))):
            if i < len(date_cells):
                row, col = date_cells[i]
                date_str = current_date.strftime("%d.%m")
                cell = worksheet.cell(row=row, column=col)
                cell.value = date_str

                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10
                current_date += timedelta(days=7)

        for i, student in enumerate(group.students):
            if i < len(student_cells):
                student_row, student_col = student_cells[i]

                grades = await Grade.filter(student_id=student.id).prefetch_related('control_work')

                for grade in grades:

                    control_work = grade.control_work

                    if control_work.discipline_id == discipline_id:

                        week = control_work.week - 1

                        if week < len(date_cells):
                            _, date_col = date_cells[week]
                            cell = worksheet.cell(row=student_row, column=date_col)
                            cell.value = grade.score

                            cell.alignment = Alignment(horizontal='center', vertical='center')

        for i in range(len(student_cells)):
            if i < len(group.students):
                student_row, _ = student_cells[i]
                for j in range(len(date_cells)):
                    _, date_col = date_cells[j]
                    cell = worksheet.cell(row=student_row, column=date_col)
                    if cell.value is None:
                        cell.value = "-"

                        cell.alignment = Alignment(horizontal='center', vertical='center')

    async def fill_student_list(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, group_id: int) -> None:
        """
        Заполняет список студентов группы в Excel

        Args:
            worksheet: Лист Excel
            group_id: ID группы
        """

        group = await Group.get(id=group_id).prefetch_related('students')

        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_value = cell.value

                if cell_value and isinstance(cell_value, str):

                    if 'S' in cell_value:
                        cell.value = cell_value.replace('S', group.code)

                    if cell_value not in ['M', 'N']:
                        continue

        data_rows = []
        student_col = None
        group_col = None
        email_col = None

        for row in range(1, worksheet.max_row + 1):
            row_has_markers = False

            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value

                if cell_value == 'M':
                    student_col = col
                    row_has_markers = True
                elif cell_value == 'N':
                    group_col = col
                    row_has_markers = True

                elif isinstance(cell_value, str) and 'email' in cell_value.lower():
                    email_col = col

            if row_has_markers:
                data_rows.append(row)

        if not data_rows:
            header_row = None

            for row in range(1, min(5, worksheet.max_row + 1)):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value

                    if isinstance(cell_value, str):
                        if 'студент' in cell_value.lower():
                            student_col = col
                            header_row = row
                        elif 'группа' in cell_value.lower():
                            group_col = col
                            header_row = row
                        elif 'email' in cell_value.lower():
                            email_col = col
                            header_row = row

                if header_row:

                    for i in range(len(group.students)):
                        data_rows.append(header_row + 1 + i)
                    break

        if not data_rows:
            start_row = 2
            student_col = student_col or 2
            group_col = group_col or 1
            email_col = email_col or 3

            for i in range(len(group.students)):
                data_rows.append(start_row + i)

        for i, student in enumerate(group.students):
            if i < len(data_rows):
                row = data_rows[i]

                if student_col:
                    cell = worksheet.cell(row=row, column=student_col)

                    if cell.value == 'M':
                        cell.value = student.full_name

                    else:
                        cell.value = student.full_name
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                if group_col:
                    cell = worksheet.cell(row=row, column=group_col)

                    if cell.value == 'N':
                        cell.value = group.code

                    else:
                        cell.value = group.code
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                if email_col:
                    cell = worksheet.cell(row=row, column=email_col)
                    cell.value = student.email
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    async def process_teacher_schedule_xlsx(self, document: openpyxl.Workbook, teacher_id: int,
                                            day_of_week: Optional[str] = None) -> openpyxl.Workbook:
        """
        Специальная обработка для расписания преподавателя в Excel

        Args:
            document: объект книги Excel
            teacher_id: ID преподавателя
            day_of_week: день недели (если не указан, берется всё расписание)

        Returns:
            Обработанный документ Excel
        """

        teacher = await Teacher.get(id=teacher_id)

        schedule_query = ScheduleItem.filter(teacher_id=teacher_id).prefetch_related(
            'time_slot', 'discipline', 'group', 'classroom'
        )

        if day_of_week:
            schedule_query = schedule_query.filter(day_of_week=day_of_week)

        schedule_items = await schedule_query.order_by('day_of_week', 'time_slot__number')

        schedule_by_day = {}
        for item in schedule_items:
            day = item.day_of_week.value
            if day not in schedule_by_day:
                schedule_by_day[day] = []
            schedule_by_day[day].append(item)

        if len(document.worksheets) == 0:
            worksheet = document.create_sheet("Расписание")
        else:
            worksheet = document.worksheets[0]

        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).value = None

        worksheet.cell(row=1, column=1).value = f"Расписание преподавателя: {teacher.full_name}"
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

        headers = ["День недели", "№ пары", "Время", "Дисциплина", "Группа", "Аудитория"]
        for i, header in enumerate(headers):
            cell = worksheet.cell(row=3, column=i + 1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = 15

        row_index = 4

        if schedule_items:
            for day, items in schedule_by_day.items():

                worksheet.cell(row=row_index, column=1).value = day
                worksheet.cell(row=row_index, column=1).font = Font(bold=True)

                for item in items:
                    worksheet.cell(row=row_index, column=2).value = item.time_slot.number
                    worksheet.cell(row=row_index,
                                   column=3).value = f"{item.time_slot.start_time}-{item.time_slot.end_time}"
                    worksheet.cell(row=row_index, column=4).value = item.discipline.name
                    worksheet.cell(row=row_index, column=5).value = item.group.code
                    worksheet.cell(row=row_index,
                                   column=6).value = item.classroom.name if item.classroom else "Нет аудитории"

                    for col in range(1, 7):
                        worksheet.cell(row=row_index, column=col).alignment = Alignment(horizontal='center',
                                                                                        vertical='center')

                    row_index += 1

                row_index += 1
        else:

            worksheet.cell(row=4, column=1).value = "У преподавателя нет занятий в расписании."
            worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)

        return document

    async def process_classroom_schedule_xlsx(self, document: openpyxl.Workbook, classroom_id: int,
                                              day_of_week: Optional[str] = None) -> openpyxl.Workbook:
        """
        Специальная обработка для расписания аудитории в Excel

        Args:
            document: объект книги Excel
            classroom_id: ID аудитории
            day_of_week: день недели (если не указан, берется всё расписание)

        Returns:
            Обработанный документ Excel
        """

        classroom = await Classroom.get(id=classroom_id)

        schedule_query = ScheduleItem.filter(classroom_id=classroom_id).prefetch_related(
            'time_slot', 'discipline', 'group', 'teacher'
        )

        if day_of_week:
            schedule_query = schedule_query.filter(day_of_week=day_of_week)

        schedule_items = await schedule_query.order_by('day_of_week', 'time_slot__number')

        schedule_by_day = {}
        for item in schedule_items:
            day = item.day_of_week.value
            if day not in schedule_by_day:
                schedule_by_day[day] = []
            schedule_by_day[day].append(item)

        if len(document.worksheets) == 0:
            worksheet = document.create_sheet("Загруженность аудитории")
        else:
            worksheet = document.worksheets[0]

        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).value = None

        worksheet.cell(row=1, column=1).value = f"Загруженность аудитории: {classroom.name}"
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

        headers = ["День недели", "№ пары", "Время", "Дисциплина", "Группа", "Преподаватель"]
        for i, header in enumerate(headers):
            cell = worksheet.cell(row=3, column=i + 1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = 15

        row_index = 4

        if schedule_items:
            for day, items in schedule_by_day.items():

                worksheet.cell(row=row_index, column=1).value = day
                worksheet.cell(row=row_index, column=1).font = Font(bold=True)

                for item in items:
                    worksheet.cell(row=row_index, column=2).value = item.time_slot.number
                    worksheet.cell(row=row_index,
                                   column=3).value = f"{item.time_slot.start_time}-{item.time_slot.end_time}"
                    worksheet.cell(row=row_index, column=4).value = item.discipline.name
                    worksheet.cell(row=row_index, column=5).value = item.group.code
                    worksheet.cell(row=row_index, column=6).value = item.teacher.full_name

                    for col in range(1, 7):
                        worksheet.cell(row=row_index, column=col).alignment = Alignment(horizontal='center',
                                                                                        vertical='center')

                    row_index += 1

                row_index += 1
        else:

            worksheet.cell(row=4, column=1).value = "В аудитории нет занятий в расписании."
            worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)

        return document

    async def process_document(self, document: openpyxl.Workbook, params: Dict[str, Any]) -> openpyxl.Workbook:
        """
        Обрабатывает XLSX-документ, заменяя плейсхолдеры на данные

        Args:
            document: Объект книги Excel
            params: Параметры для обработки документа

        Returns:
            Обработанный документ Excel
        """
        self.replacements = {}

        template_id = params.get('template_id')
        group_id = params.get('group_id')
        student_id = params.get('student_id')
        teacher_id = params.get('teacher_id')
        discipline_id = params.get('discipline_id')
        start_date_str = params.get('start_date')
        day_of_week = params.get('day_of_week')
        classroom_id = params.get('classroom_id')

        start_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            except (ValueError, TypeError):
                start_date = datetime.now()

        template = await self.get_template(template_id)

        is_journal = False
        is_student_list = False
        is_teacher_schedule = False
        is_classroom_schedule = False

        if 'журнал' in template.name.lower() or 'journal' in template.name.lower():
            is_journal = True
        elif 'список' in template.name.lower() or 'list' in template.name.lower() or 'студент' in template.name.lower():
            is_student_list = True
        elif 'расписание_преподавателя' in template.name.lower() or 'teacher_schedule' in template.name.lower():
            is_teacher_schedule = True
        elif 'загруженность_аудитории' in template.name.lower() or 'classroom_schedule' in template.name.lower():
            is_classroom_schedule = True

        if not is_journal and not is_student_list and not is_teacher_schedule and not is_classroom_schedule:

            for sheet_name in document.sheetnames:
                worksheet = document[sheet_name]
                for row in range(1, min(10, worksheet.max_row + 1)):
                    for col in range(1, min(10, worksheet.max_column + 1)):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value == 'd':
                            is_journal = True
                            break
                        elif cell_value in ['M', 'N']:
                            is_student_list = True
                            break

        if group_id:
            group = await Group.get(id=group_id).prefetch_related('students')
            self.replacements.update({
                "group": group.code,
                "course": group.course,
                "S": group.code
            })

        if student_id:
            student = await Student.get(id=student_id).prefetch_related('group')
            self.replacements.update({
                "student_name": student.full_name,
                "student_email": student.email,
                "G": student.full_name,
                "M": student.full_name
            })

            if student.group:
                self.replacements.update({
                    "group": student.group.code,
                    "S": student.group.code,
                })

        if teacher_id:
            teacher = await Teacher.get(id=teacher_id)
            self.replacements.update({
                "teacher_name": teacher.full_name,
                "teacher_email": teacher.email,
                "P": teacher.full_name,
                "B": teacher.full_name
            })

        if discipline_id:
            discipline = await Discipline.get(id=discipline_id)
            self.replacements.update({
                "discipline": discipline.name,
                "N": discipline.name,
            })

        if is_teacher_schedule and teacher_id:

            document = await self.process_teacher_schedule_xlsx(document, teacher_id, day_of_week)
        elif is_classroom_schedule and classroom_id:

            document = await self.process_classroom_schedule_xlsx(document, classroom_id, day_of_week)
        else:

            for sheet_name in document.sheetnames:
                worksheet = document[sheet_name]

                if is_journal and group_id and discipline_id:

                    await self.fill_grades_journal(worksheet, group_id, discipline_id, start_date)
                elif is_student_list and group_id:

                    await self.fill_student_list(worksheet, group_id)
                else:

                    for row in range(1, worksheet.max_row + 1):
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)

                            if cell.value and isinstance(cell.value, str) and cell.value not in ['M', 'N', 'd']:

                                if any(placeholder in cell.value for placeholder in self.replacements.keys()) or any(
                                        "{{" + placeholder + "}}" in cell.value for placeholder in
                                        self.replacements.keys()):
                                    cell.value = await self.replace_placeholders_in_text(cell.value, self.replacements)

        return document
